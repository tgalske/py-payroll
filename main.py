import openpyxl
from employee import Employee
from worked_job import WorkedJob

workbook = None
current_sheet = None
current_row = 2
weekly_earnings = {}  # Key:Value mapping of employee IDs to employee objects to store earnings

WORKBOOK_NAME = 'office_hours.xlsx'
SHEET_NAME = '4.22-4.29'
EMP_ID_COLUMN = 1
EMP_FIRSTNAME_COLUMN = 2
EMP_LASTNAME_COLUMN = 3
JOB_NAME_COLUMN = 4
JOB_PAY_COLUMN = 5
HOURS_COLUMN = 6


def row_is_empty(row_number):
    return current_sheet.cell(row=row_number, column=1).value is None


# returns true if the earnings map contains an entry for an employee based on an employee ID
def map_contains_employee(employee_id):
    return employee_id in weekly_earnings


# Main program

# Load workbook and active sheet
workbook = openpyxl.load_workbook(WORKBOOK_NAME)
current_sheet = workbook[SHEET_NAME]

# Loop through all rows (that contain information in column 1)
while not row_is_empty(current_row):
    employee_id = current_sheet.cell(row=current_row, column=EMP_ID_COLUMN).value
    job_name = current_sheet.cell(row=current_row, column=JOB_NAME_COLUMN).value
    job_pay = current_sheet.cell(row=current_row, column=JOB_PAY_COLUMN).value
    hours = current_sheet.cell(row=current_row, column=HOURS_COLUMN).value
    current_worked_job = WorkedJob(job_name, job_pay, hours)

    if map_contains_employee(employee_id):
        # the employee already exists in the map, so add the new worked job
        current_employee = weekly_earnings[employee_id]
    else:
        # the employee does not exist in the map yet, so create a new employee object
        firstname = current_sheet.cell(row=current_row, column=EMP_FIRSTNAME_COLUMN).value
        lastname = current_sheet.cell(row=current_row, column=EMP_LASTNAME_COLUMN).value
        current_employee = Employee(employee_id, firstname, lastname)

        # add the employee to the weekly earnings map
        weekly_earnings[employee_id] = current_employee

    # add the current job to the employee's list of worked jobs
    current_employee.add_worked_job(current_worked_job)

    # go to the next row
    current_row += 1

# print out employee information for each employee contained in the map
for employee_id in weekly_earnings:
    print(weekly_earnings[employee_id].to_string())
